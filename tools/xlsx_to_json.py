#!/usr/bin/env python3
import json
from pathlib import Path
import openpyxl

def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s != "" else None
    return v

def main():
    repo = Path(__file__).resolve().parents[1]
    xlsx_path = repo / "NPCs.xlsx"
    json_path = repo / "NPCs.json"

    if not xlsx_path.exists():
        raise SystemExit(f"NPCs.xlsx not found at {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["NPCs"] if "NPCs" in wb.sheetnames else wb[wb.sheetnames[0]]

    headers = []
    for cell in ws[1]:
        h = clean(cell.value)
        headers.append(str(h) if h is not None else "")

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or v == "" for v in row):
            continue

        obj = {}
        for k, v in zip(headers, row):
            if not k:
                continue
            v = clean(v)

            if isinstance(v, str):
                lv = v.lower()
                if lv in ("true", "yes", "y", "1"):
                    v = True
                elif lv in ("false", "no", "n", "0"):
                    v = False

            obj[k] = v

        name = obj.get("Name")
        if not isinstance(name, str) or not name.strip():
            continue

        out.append(obj)

    json_path.write_text(
        json.dumps(out, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    print(f"Wrote {len(out)} NPCs -> {json_path}")

if __name__ == "__main__":
    main()
